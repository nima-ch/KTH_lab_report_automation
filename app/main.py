from kivy.app import App
from kivy.properties import BooleanProperty, StringProperty
from kivy.uix.boxlayout import BoxLayout
from kivy.core.window import Window
from kivy.uix.boxlayout import BoxLayout
from openpyxl import Workbook
import os
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side


class MainUi(BoxLayout):

    main_message = StringProperty("Drop your folder here")
    report_found = BooleanProperty(False)

    # Your BooleanProperty variables
    retention_time_checked = BooleanProperty(True)
    area_checked = BooleanProperty(True)
    amount_checked = BooleanProperty(True)
    amount_area_checked = BooleanProperty(True)

    df_dict = {}
    root_dir = "./"

    # Define the initial list of keys
    check_box_list = ['Retention Time', 'Area', 'Amount', 'Amount/Area']


    def __init__(self, **kwargs):
        super(MainUi, self).__init__(**kwargs)
        Window.bind(on_drop_file=self._on_file_drop)
        
    def update_checkbox_list(self, checkbox_label, value):
        if value:
            if checkbox_label not in self.check_box_list:
                self.check_box_list.append(checkbox_label)
        else:
            if checkbox_label in self.check_box_list:
                self.check_box_list.remove(checkbox_label)
        print(f"Updated checkbox list: {self.check_box_list}")

    def extract_data_from_report(self, file_path):
        try:
            with open(file_path, 'r', encoding="utf-16-le") as file:
                lines = file.readlines()
                content = ''.join(lines)

                # External Standard Report
                words_to_check = {"RetTime", "Sig", "Type",
                                  "Area", "Amt/Area", "Amount", "Grp", "Name"}

                # Check if exactly one line contains the set of words_to_check, allowing for "Sig" to be optional
                if sum(set(words_to_check).difference(set(line.split())) in [set(), {"Sig"}] for line in lines) == 1:
                    # Extract sample name
                    sample_name_match = re.search(
                        r'Sample Name: (\S+)', content)
                    sample_name = sample_name_match.group(
                        1) if sample_name_match else "Unknown"

                    # Regex patterns for different table formats
                    pattern1 = r'(\d+\.\d{3})\s+(\d+)\s+([A-Za-z]+)?\s+(-|[\de\.-]+)\s+(-|[\de\.-]+)\s+(-|[\de\.-]+)\s+([A-Za-z0-9-_, ]+)'
                    pattern2 = r'(\d+\.\d{3})\s+([A-Za-z]+)?\s+(-|[\de\.-]+)\s+(-|[\de\.-]+)\s+(-|[\de\.-]+)\s+([A-Za-z0-9-_, ]+)'

                    # Use the first pattern
                    data_matches = re.findall(pattern1, content)
                    if not data_matches:
                        # If no match, try the second pattern
                        data_matches = re.findall(pattern2, content)
                    data = {
                        'Sample Name': sample_name,
                        'Entries': []
                    }
                    for match in data_matches:
                        if len(match) == 7:
                            ret_time, signal, type_, area, amt_area, amt, grp_name = match
                        else:
                            ret_time, type_, area, amt_area, amt, grp_name = match
                            signal = None

                        entry = {
                            'Retention Time': float(ret_time),
                            'Signal': None if signal is None else int(signal),
                            'Type': type_.strip() if type_ else None,
                            'Area': 0 if area == '-' else float(area),
                            'Amount/Area': 0 if amt_area == '-' else float(amt_area),
                            'Amount': 0 if amt == '-' else float(amt),
                            'Standard': grp_name.strip()
                        }
                        data['Entries'].append(entry)
                    return data
                else:
                    return None
        except Exception as e:
            print(f"An error occurred: {e}")
            return None

    def _on_file_drop(self, window, file_path, x, y):

        self.main_message = file_path.decode("UTF-8")
        self.root_dir = file_path.decode("UTF-8")
        files_path_list = []
        data = []
        self.df_dict = {}

        for folder, subfolders, files in os.walk(self.root_dir):
            for f in files:
                if f.endswith("Report.TXT"):
                    files_path_list.append(os.path.join(folder, f))
        if len(files_path_list) == 0:
            self.main_message = f"NO REPORT HAS BEEN FOUND IN \n{self.root_dir}"
        else:
            for file in files_path_list:
                data_dict = self.extract_data_from_report(file)
                data.append(data_dict)
            self.main_message = f"{len(files_path_list)} folders contain report files of which {len([element for element in data if element is not None])} can be converted to excel."
            self.report_found = True
        for smp in data:
            if smp:
                self.df_dict.update(
                    {smp['Sample Name']: pd.DataFrame(smp['Entries'])})

    def get_df_value(self, col):
        data_list = []
        for sample_name, df in self.df_dict.items():
            df_filtered = df.dropna(subset=['Standard', col])
            area_dict = df_filtered.set_index('Standard')[col].to_dict()
            area_dict['Sample Name'] = sample_name
            data_list.append(area_dict)

        result_df = pd.DataFrame(data_list)

        cols = ['Sample Name'] + \
            [col for col in result_df.columns if col != 'Sample Name']
        return result_df[cols]

    def on_create_click(self):
        # Define the list of keys
        keys = self.check_box_list

        # Initialize an empty dictionary
        df_dictionary = {}

        # Populate the dictionary with DataFrames
        for key in keys:
            if key == 'Retention Time':
                ret_df = self.get_df_value(col=key)
                ret_df.loc['Average'] = ret_df.mean(numeric_only=True)
                ret_df.at['Average', 'Sample Name'] = 'Average'
                df_dictionary[key] = ret_df
            else:
                df_dictionary[key] = self.get_df_value(col=key)

        # Initialize a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Define the center alignment
        center_aligned = Alignment(horizontal='center', vertical='center')

        # Initialize the row number
        row_num = 1

        # Define a list of border colors
        border_colors = ['2f7c47', '3b8eea', 'ff6034', '1f1f1f']

        for idx, (key, df) in enumerate(df_dictionary.items()):
            # Define the border with color
            border = Border(left=Side(border_style='medium', color=border_colors[idx]),
                            right=Side(border_style='medium',
                                       color=border_colors[idx]),
                            top=Side(border_style='medium',
                                     color=border_colors[idx]),
                            bottom=Side(border_style='medium', color=border_colors[idx]))

            # Add table header
            header_cell = ws.cell(row=row_num, column=1, value=key)
            header_cell.alignment = center_aligned

            row_num += 1  # Move to the next row for column headers and data

            # Write the dataframe to Excel
            for r_idx, row in enumerate(df.values):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=row_num + r_idx,
                                   column=c_idx + 1, value=value)
                    cell.alignment = center_aligned
                    cell.border = border

            # Write the column headers
            for c_idx, col in enumerate(df.columns.values):
                cell = ws.cell(row=row_num, column=c_idx + 1, value=col)
                cell.alignment = center_aligned
                cell.border = border

            row_num += df.shape[0] + 2  # Skip one row before the next table

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save the workbook
        wb.save(f"{self.root_dir}/final_report.xlsx")
        self.main_message = "Excle file successfuly created in the same directory"


class XlsxApp(App):
    pass


if __name__ == '__main__':
    XlsxApp().run()
